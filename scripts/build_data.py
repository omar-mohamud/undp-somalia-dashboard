"""Extract pipeline_and_programme_v2.xlsx into data/data.json.

Re-run whenever the workbook changes.
"""

import json
import re
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
WB_PATH = ROOT / "pipeline_and_programme_v2.xlsx"
OUT_PATH = ROOT / "data" / "data.json"

# cp1252 leftovers that pandas/openpyxl can leak into strings
CHAR_FIXES = {
    "‘": "'", "’": "'", "‚": "'", "‛": "'",
    "“": '"', "”": '"', "„": '"', "‟": '"',
    "–": "-", "—": "-", "−": "-",
    "…": "...",
    "\xa0": " ",
    "\x92": "'", "\x91": "'",
    "\x93": '"', "\x94": '"',
    "\x96": "-", "\x97": "-",
}


def fix_str(s: str) -> str:
    for a, b in CHAR_FIXES.items():
        s = s.replace(a, b)
    return re.sub(r"\s+", " ", s).strip()


def clean(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        return fix_str(v)
    return v


def rows(ws):
    it = ws.iter_rows(values_only=True)
    headers = [h for h in next(it)]
    out = []
    for r in it:
        if all(c is None for c in r):
            continue
        out.append({headers[i]: clean(r[i]) for i in range(len(headers))})
    return out


def fnum(x):
    try:
        return round(float(x), 2) if x is not None else 0
    except Exception:
        return 0


def main():
    wb = load_workbook(WB_PATH, data_only=True)

    donors_raw = rows(wb["dim_Donor"])
    portfolios_raw = rows(wb["dim_Portfolio"])
    stages_raw = rows(wb["dim_Pipeline_Stage"])
    pipeline_raw = rows(wb["fact_Pipeline"])
    cash_raw = rows(wb["fact_CashReceived"])
    delivery_raw = rows(wb["fact_Delivery"])
    quarterly_raw = rows(wb["fact_QuarterlyTargets"])
    kpis_raw = rows(wb["ref_PartnershipKPIs"])

    donor_by_id = {d["DonorID"]: d for d in donors_raw}
    portfolio_by_id = {p["PortfolioID"]: p for p in portfolios_raw}

    portfolios = [
        {
            "id": p["PortfolioID"],
            "code": p["Portfolio_Code"],
            "name": p["Portfolio_Full_Name"],
            "annualTarget": fnum(p["Annual_Target_USD_2026"]),
        }
        for p in portfolios_raw
        if p["PortfolioID"] != "P99"
    ]

    stages = [
        {
            "code": s["Stage_Code"],
            "label": re.sub(r"^[A-D]\s*-\s*", "", s["Stage_Label"]),
            "fullLabel": s["Stage_Label"],
            "probability": fnum(s["Probability"]),
            "order": int(s["Stage_Order"]),
            "description": s["Description"],
        }
        for s in stages_raw
    ]

    opportunities = []
    for o in pipeline_raw:
        donor = donor_by_id.get(o["DonorID"], {})
        portfolio = portfolio_by_id.get(o["PortfolioID"], {})
        opportunities.append(
            {
                "id": o["OpportunityID"],
                "donorId": o["DonorID"],
                "donorName": donor.get("Donor_Canonical_Name") or o["Donor"],
                "donorAlias": o["Donor"],
                "donorType": donor.get("Donor_Type") or "Other",
                "donorCountry": donor.get("Donor_Country"),
                "portfolioId": o["PortfolioID"],
                "portfolioCode": portfolio.get("Portfolio_Code") or o["Pillar"],
                "portfolioName": portfolio.get("Portfolio_Full_Name"),
                "programme": o["Programme"],
                "stage": o["Pipeline_Stage"],
                "stageLabel": re.sub(
                    r"^[A-D]\s*-\s*", "", o["Pipeline_Stage_Label"] or ""
                ),
                "pipelineTotal": fnum(o["Pipeline_Total_USD"]),
                "target2026": fnum(o["Target_2026_USD"]),
                "target2026Adj": fnum(o["Target_2026_Adj_USD"]),
                "target2027": fnum(o["Target_2027_USD"]),
                "target2027Adj": fnum(o["Target_2027_Adj_USD"]),
                "reportDate": o["ReportDate"],
            }
        )

    donor_agg = {}
    for o in opportunities:
        d = donor_agg.setdefault(
            o["donorId"],
            {
                "opportunities": 0,
                "pipelineTotal": 0,
                "target2026Adj": 0,
                "target2027Adj": 0,
                "stages": {},
            },
        )
        d["opportunities"] += 1
        d["pipelineTotal"] += o["pipelineTotal"]
        d["target2026Adj"] += o["target2026Adj"]
        d["target2027Adj"] += o["target2027Adj"]
        d["stages"][o["stage"]] = d["stages"].get(o["stage"], 0) + 1

    cash_by_donor = {}
    for c in cash_raw:
        cash_by_donor.setdefault(c["DonorID"], 0)
        cash_by_donor[c["DonorID"]] += fnum(c["Cash_Received_USD"])

    donors = []
    for d in donors_raw:
        if d["DonorID"] == "D999":
            continue
        agg = donor_agg.get(d["DonorID"], {})
        donors.append(
            {
                "id": d["DonorID"],
                "name": d["Donor_Canonical_Name"],
                "type": d["Donor_Type"],
                "country": d["Donor_Country"],
                "opportunityCount": agg.get("opportunities", 0),
                "pipelineTotal": round(agg.get("pipelineTotal", 0), 2),
                "target2026Adj": round(agg.get("target2026Adj", 0), 2),
                "target2027Adj": round(agg.get("target2027Adj", 0), 2),
                "cashReceived": round(cash_by_donor.get(d["DonorID"], 0), 2),
                "stages": agg.get("stages", {}),
            }
        )

    cash = []
    for c in cash_raw:
        cash.append(
            {
                "id": c["CashID"],
                "donorId": c["DonorID"],
                "portfolioId": c["PortfolioID"],
                "donorAlias": c["Donor"],
                "donorCountry": c["Donor_Country"],
                "programme": c["Programme"],
                "agreementAmount": fnum(c["Agreement_Amount_USD"]),
                "cashReceived": fnum(c["Cash_Received_USD"]),
                "date": c["Date"],
                "year": c["Year_Num"],
                "quarter": c["Quarter"],
                "status": c["Status"],
                "fundType": c["Fund_Type"],
            }
        )

    delivery_by_portfolio = []
    for d in delivery_raw:
        delivery_by_portfolio.append(
            {
                "portfolioId": d["PortfolioID"],
                "portfolio": (d["Portfolio"] or "").strip() or None,
                "year": d["Year"],
                "annualTarget": fnum(d["Annual_Target_USD"]),
                "totalBudget": fnum(d["Total_Budget_USD"]),
                "totalExpenditure": fnum(d["Total_Expenditure_USD"]),
                "budgetBalance": fnum(d["Budget_Balance_USD"]),
                "pctDeliveryBudget": fnum(d["Pct_Delivery_Budget"]),
                "pctDeliveryAnnualTarget": fnum(d["Pct_Delivery_Annual_Target"]),
            }
        )

    quarterly_targets = []
    for q in quarterly_raw:
        quarterly_targets.append(
            {
                "portfolioId": q["PortfolioID"],
                "portfolio": q["Portfolio"],
                "q1Actual": fnum(q["Q1_Actual_USD"]),
                "q2Target": fnum(q["Q2_Target_USD"]),
                "q3Target": fnum(q["Q3_Target_USD"]),
                "q4Target": fnum(q["Q4_Target_USD"]),
                "annualTarget": fnum(q["Annual_Target_USD"]),
            }
        )

    kpis = {}
    for k in kpis_raw:
        cat = k["Category"]
        kpis.setdefault(cat, []).append(
            {
                "indicator": k["Indicator"],
                "value": k["Value"],
                "unit": k["Unit"],
            }
        )

    by_stage = {}
    for o in opportunities:
        s = by_stage.setdefault(
            o["stage"], {"stage": o["stage"], "value": 0, "count": 0}
        )
        s["value"] += o["pipelineTotal"]
        s["count"] += 1
    stage_label_map = {s["code"]: s["label"] for s in stages}
    stage_order_map = {s["code"]: s["order"] for s in stages}
    pipeline_by_stage = []
    for code, agg in by_stage.items():
        pipeline_by_stage.append(
            {
                "stage": code,
                "label": stage_label_map.get(code, code),
                "value": round(agg["value"], 2),
                "count": agg["count"],
                "order": stage_order_map.get(code, 99),
            }
        )
    pipeline_by_stage.sort(key=lambda r: r["order"])

    by_type = {}
    for o in opportunities:
        s = by_type.setdefault(
            o["donorType"], {"type": o["donorType"], "value": 0, "count": 0, "donors": set()}
        )
        s["value"] += o["pipelineTotal"]
        s["count"] += 1
        s["donors"].add(o["donorId"])
    pipeline_by_donor_type = [
        {
            "type": t,
            "value": round(v["value"], 2),
            "count": v["count"],
            "donorCount": len(v["donors"]),
        }
        for t, v in by_type.items()
    ]
    pipeline_by_donor_type.sort(key=lambda r: -r["value"])

    by_portfolio = {}
    for o in opportunities:
        s = by_portfolio.setdefault(
            o["portfolioCode"],
            {"portfolioCode": o["portfolioCode"], "value": 0, "count": 0},
        )
        s["value"] += o["pipelineTotal"]
        s["count"] += 1
    pipeline_by_portfolio = list(by_portfolio.values())
    for p in pipeline_by_portfolio:
        p["value"] = round(p["value"], 2)
    pipeline_by_portfolio.sort(key=lambda r: -r["value"])

    type_counts = {}
    for d in donors:
        type_counts.setdefault(d["type"], 0)
        type_counts[d["type"]] += 1

    pipeline_total = round(sum(o["pipelineTotal"] for o in opportunities), 2)
    target_2026_adj = round(sum(o["target2026Adj"] for o in opportunities), 2)
    target_2027_adj = round(sum(o["target2027Adj"] for o in opportunities), 2)
    cash_received_2025 = round(
        sum(c["cashReceived"] for c in cash if c["year"] == 2025), 2
    )
    cash_received_2026 = round(
        sum(c["cashReceived"] for c in cash if c["year"] == 2026), 2
    )

    delivery_total_target = round(sum(d["annualTarget"] for d in delivery_by_portfolio), 2)
    delivery_total_budget = round(sum(d["totalBudget"] for d in delivery_by_portfolio), 2)
    delivery_total_exp = round(sum(d["totalExpenditure"] for d in delivery_by_portfolio), 2)

    out = {
        "meta": {
            "asOfDate": "2026-03-31",
            "lastUpdated": "2026-04-30",
            "reportDate": "2026-04-28",
            "schemaVersion": "2.0",
        },
        "totals": {
            "pipelineTotal": pipeline_total,
            "target2026Adj": target_2026_adj,
            "target2027Adj": target_2027_adj,
            "cashReceived2025": cash_received_2025,
            "cashReceived2026": cash_received_2026,
            "opportunityCount": len(opportunities),
            "donorCountWithOpps": len({o["donorId"] for o in opportunities}),
            "donorCountTotal": len(donors),
            "deliveryAnnualTarget": delivery_total_target,
            "deliveryTotalBudget": delivery_total_budget,
            "deliveryTotalExpenditure": delivery_total_exp,
            "deliveryPctVsTarget": round(delivery_total_exp / delivery_total_target, 4)
            if delivery_total_target
            else 0,
            "deliveryPctVsBudget": round(delivery_total_exp / delivery_total_budget, 4)
            if delivery_total_budget
            else 0,
        },
        "donors": donors,
        "donorTypeCounts": [
            {"type": t, "count": c}
            for t, c in sorted(type_counts.items(), key=lambda r: -r[1])
        ],
        "portfolios": portfolios,
        "stages": stages,
        "opportunities": opportunities,
        "cash": cash,
        "deliveryByPortfolio": delivery_by_portfolio,
        "quarterlyTargets": quarterly_targets,
        "kpis": kpis,
        "pipelineByStage": pipeline_by_stage,
        "pipelineByDonorType": pipeline_by_donor_type,
        "pipelineByPortfolio": pipeline_by_portfolio,
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(
        json.dumps(out, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    print(
        f"Wrote {OUT_PATH} - {len(opportunities)} opps, "
        f"{len(donors)} donors, {len(cash)} cash records"
    )


if __name__ == "__main__":
    main()
